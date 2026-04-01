import openpyxl
import json
import sys

print("DEBUG: Script started")
file_path = 'EFETIVO.xlsx'
try:
    print(f"DEBUG: Loading workbook {file_path} (read_only=True)...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print("DEBUG: Workbook loaded. Getting active sheet...")
    sheet = wb.active
    print(f"DEBUG: Active sheet: {sheet.title}. Reading headers...")
    
    # Get only the first row (headers)
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    print(f"DEBUG: Headers found: {headers}")
    
    print("DEBUG: Reading first 5 data rows...")
    data_rows = []
    # Skip header
    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
        data_rows.append(dict(zip(headers, row)))
    
    res = {
        'columns': headers,
        'head': data_rows
    }
    print("RESULT_START")
    print(json.dumps(res, indent=2, ensure_ascii=False))
    print("RESULT_END")
except Exception as e:
    print(f"DEBUG: Error: {e}")
finally:
    print("DEBUG: Script finished")

sys.stdout.flush()
